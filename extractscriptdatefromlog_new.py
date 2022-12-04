from distutils.log import info
import numpy as np
import os
from openpyxl import load_workbook
from openpyxl import Workbook
import tkinter as tk
from tkinter import filedialog
from tkinter import Button
from datetime import datetime
import time

timestr = time.strftime("%Y%m%d-%H%M%S")
analysis_filename = timestr + str("_Alpha_run_log_analysis.xlsx")


def getdirectory():
    directory = filedialog.askdirectory()
    print(directory)
    # for subdir, dirs, files in os.walk(directory):
    #     # for file in files:
    #     if "application" in subdir:
    #         Module_SN = subdir.partition("v")[0]
    #         Module_SN = Module_SN.partition("Logs")[2]
    #         print(subdir)
    #         print(Module_SN)
    array = main_func(directory)
    write_data(array, analysis_filename, directory)
    root.destroy()


def main_func(directory):
    filerun = []
    for subdir, dirs, files in os.walk(directory):
        # for file in files:
        if "application" in subdir:
            Module_SN = subdir.partition("v")[0]
            Module_SN = Module_SN.partition("Logs")[2]
            # print(subdir)
            # print(Module_SN)

            PH_version = subdir.partition("v")[2]
            PH_version = PH_version.partition("application")[0]
            # print(subdir)
            # print(PH_version)
            for filename in os.listdir(subdir):
                if filename.endswith(".log"):
                    fn = open(os.path.join(subdir, filename), "r")
                    data = fn.read()
                    fn.close()
                    datalog = data.split("\n")
                    information_list = [""] * 9
                    for i in range(len(datalog)):
                        line = datalog[i]
                        if "example_scripts" in line:
                            filerun = filerun + [information_list]
                            information_list = [""] * 12
                            information_list[10] = PH_version
                            information_list[11] = Module_SN
                            script = line.partition("example_scripts")[2]
                            script = script.partition(" to system path")[0]
                            information_list[1] = script
                        if "Starting external script" in line:
                            first_log_line = i
                            date_started = line.partition("Scripting")[0]
                            if date_started:
                                time_started = datetime.strptime(
                                    date_started, "%Y-%m-%d %H:%M:%S,%f "
                                )

                                actual_date = datetime.strptime(
                                    time_started.strftime("%Y-%m-%d"), "%Y-%m-%d"
                                ).date()

                                information_list[0] = filename
                                information_list[2] = actual_date
                                information_list[3] = time_started.strftime("%H:%M:%S")
                        if "Finished running script" in line:
                            last_log_line = i
                            date_finished = line.partition("Scripting")[0]
                            if date_finished:
                                time_finished = datetime.strptime(
                                    date_finished, "%Y-%m-%d %H:%M:%S,%f "
                                )
                                script_duration = time_finished - time_started
                                information_list[4] = date_finished
                                information_list[5] = script_duration

                            if "User stoped script" in datalog[i - 2]:
                                information_list[6] = "NO"
                            else:
                                information_list[6] = "YES"

                            for a in range(last_log_line - first_log_line):
                                log_line = datalog[a + first_log_line]
                                if "System transaction list is full" in log_line:
                                    information_list[7] = "error_image_transfer"
                                if "STEPLOSS" in log_line:
                                    information_list[8] = "error_steploss"
                                if "Failed to acquire" in log_line:
                                    information_list[9] = "error_image_acquisition"

                    filerun = filerun + [information_list]

    return filerun


def write_data(array, filename, directory):
    # load excel file, need to have one named ZOffsetData
    filename = os.path.join(directory, filename)

    try:
        wb = load_workbook(filename=filename)
    except FileNotFoundError:
        wb = Workbook()
    ws = wb.active
    col = ws.max_column
    row = 1
    if col == 1:
        ws.cell(row, col, "Log")
        ws.cell(row, col + 1, "Script Loaded")
        ws.cell(row, col + 2, "Date Started")
        ws.cell(row, col + 3, "Time Started")
        ws.cell(row, col + 4, "Date / Time finished")
        ws.cell(row, col + 5, "Script Duration")
        ws.cell(row, col + 6, "Run Completed")
        ws.cell(row, col + 7, "Error Image Transfer")
        ws.cell(row, col + 8, "Error Steploss")
        ws.cell(row, col + 9, "Error Image Acquisition")
        ws.cell(row, col + 10, "PH Version")
        ws.cell(row, col + 11, "Alpha Module")

    # write to excel file

    row += 1
    # for log in array:
    print(array)
    for item in array:
        written = False
        index = 1
        for i in range(len(item)):
            newitem = item[i]
            if item[2] and item[1]:
                ws.cell(row, index, newitem)
                index += 1
                written = True
        if written:
            row += 1

    while True:
        try:
            wb.save(filename)
            print("saved")
            break
        except:
            print("Close Window")


# Create GUI
root = tk.Tk()
root.title("Extractor")
root.geometry("400x400")
root.configure(bg="green")
button = Button(
    root,
    text="Choose a Folder",
    command=getdirectory,
    width=12,
    height=4,
    font="Times 14 bold",
    bg="white",
    fg="#56b5fd",
)
button.place(x=142, y=142)

root.mainloop()
